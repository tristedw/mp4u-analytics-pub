import openpyxl

from ui_main import *
from PySide6.QtCore import *
from PySide6.QtGui import *
from PySide6.QtWidgets import *
from PySide6.QtCharts import *
from py_left_button import PyLeftButton
from py_push_button import PyPushButton
from py_table_widget import PyTableWidget


class MainFunctions:
    created_widgets = []
    is_file_selected = False

    def __init__(self):
        super().__init__()

        self.ui = Ui_main_window()
        self.ui.setupUi(self)

    def setup_widgets(self):
        # Toggle Button
        self.toggle_btn = PyLeftButton(
            "Hide menu",
            icon_name="icon_menu.png"
        )
        self.toggle_btn.clicked.connect(lambda: MainFunctions.toggle_button(self))

        self.orders_button = PyLeftButton(
            "Orders",
            icon_name="icon_orders.png",
            is_active=False
        )

        self.orders_button.clicked.connect(lambda: MainFunctions.show_orders(self, self.orders_button))

        self.finances_button = PyLeftButton(
            "Finances",
            icon_name="icon_finances.png",
            is_active=False
        )
        self.finances_button.clicked.connect(lambda: MainFunctions.show_finances(self, self.finances_button))

        self.employees_button = PyLeftButton(
            "Employees",
            icon_name="icon_employees.png",
            is_active=False
        )
        self.employees_button.clicked.connect(lambda: MainFunctions.show_employees(self, self.employees_button))

        self.select_file_button = PyLeftButton(
            "Select File",
            icon_name="icon_select_file.png",
            is_active=True
        )
        self.select_file_button.clicked.connect(lambda: MainFunctions.show_select_file(self, self.select_file_button))

        self.compare_file_button = PyLeftButton(
            "Select Compare File",
            icon_name="icon_compare_file.png",
            is_active=False
        )
        self.compare_file_button.clicked.connect(
            lambda: MainFunctions.show_compare_file(self, self.compare_file_button))

        # Add Widgets To Menu
        self.ui.top_menu_layout.addWidget(self.toggle_btn)
        self.ui.top_menu_layout.addWidget(self.orders_button)
        self.ui.top_menu_layout.addWidget(self.finances_button)
        self.ui.top_menu_layout.addWidget(self.employees_button)
        self.ui.bottom_menu_layout.addWidget(self.select_file_button)
        self.ui.bottom_menu_layout.addWidget(self.compare_file_button)

        # Select file page
        self.select_file_button_controller = PyPushButton(
            text="Select File",
            expanding=True,
            is_compare=False
        )
        self.select_file_button_controller.clicked.connect(lambda: self.parse_data(False))

        self.ui.file_select_layout.addWidget(self.select_file_button_controller)

        # Compare file page
        self.compare_file_button_controller = PyPushButton(
            text="Select Compare File",
            expanding=True,
            is_compare=False
        )
        self.compare_file_button_controller.clicked.connect(lambda: self.parse_data(True))

        self.ui.compare_file_layout.addWidget(self.compare_file_button_controller)

        self.ui.label_2.setText("V1.4")

    # Reset BTN Selection
    def reset_selection(self):
        for btn in self.ui.left_menu.findChildren(QPushButton):
            try:
                btn.set_active(False)
            except:
                pass

    # Buttons
    def show_orders(self, btn):
        if not MainFunctions.is_file_selected:
            MainFunctions.reset_selection(self)  # Deselect All Buttons
            self.ui.app_pages.setCurrentWidget(self.ui.page_file_select)  # Change Page
            self.select_file_button.set_active(True)  # Set Active Button
            return

        MainFunctions.reset_selection(self)  # Deselect All Buttons
        self.ui.app_pages.setCurrentWidget(self.ui.page_orders)  # Change Page
        btn.set_active(True)  # Set Active Button

    def show_finances(self, btn):
        if not MainFunctions.is_file_selected:
            MainFunctions.reset_selection(self)  # Deselect All Buttons
            self.ui.app_pages.setCurrentWidget(self.ui.page_file_select)  # Change Page
            self.select_file_button.set_active(True)  # Set Active Button
            return

        MainFunctions.reset_selection(self)  # Deselect All Buttons
        self.ui.app_pages.setCurrentWidget(self.ui.page_finances)  # Change Page
        btn.set_active(True)  # Set Active Button

    def show_employees(self, btn):
        if not MainFunctions.is_file_selected:
            MainFunctions.reset_selection(self)  # Deselect All Buttons
            self.ui.app_pages.setCurrentWidget(self.ui.page_file_select)  # Change Page
            self.select_file_button.set_active(True)  # Set Active Button
            return

        MainFunctions.reset_selection(self)  # Deselect All Buttons
        self.ui.app_pages.setCurrentWidget(self.ui.page_employees)  # Change Page
        btn.set_active(True)  # Set Active Button

    def show_select_file(self, btn):
        MainFunctions.reset_selection(self)  # Deselect All Buttons
        self.ui.app_pages.setCurrentWidget(self.ui.page_file_select)  # Change Page
        btn.set_active(True)  # Set Active Button

    def show_compare_file(self, btn):
        if not MainFunctions.is_file_selected:
            MainFunctions.reset_selection(self)  # Deselect All Buttons
            self.ui.app_pages.setCurrentWidget(self.ui.page_file_select)  # Change Page
            self.select_file_button.set_active(True)  # Set Active Button
            return
        MainFunctions.reset_selection(self)  # Deselect All Buttons
        self.ui.app_pages.setCurrentWidget(self.ui.page_compare_file_select)  # Change Page
        btn.set_active(True)  # Set Active Button

    def toggle_button(self):
        # Get menu width
        menu_width = self.ui.left_menu.width()
        self.ui.app_pages.repaint()
        self.ui.app_pages.repaint()

        # Check width
        width = 60
        if menu_width == 60:
            width = 240

        # Start animation
        self.animation = QPropertyAnimation(self.ui.left_menu, b"minimumWidth")

        self.animation.setStartValue(menu_width)
        self.animation.setEndValue(width)
        self.animation.setDuration(200)
        self.animation.setEasingCurve(QEasingCurve.InOutCirc)
        self.animation.start()

    def occupy_tabs(self):
        # Orders page
        self.orders_table = PyTableWidget()
        MainFunctions.created_widgets.append(self.orders_table)

        self.orders_pie_chart_series = QPieSeries()

        self.orders_pie_chart = QChart()
        self.orders_pie_chart.addSeries(self.orders_pie_chart_series)
        self.orders_pie_chart.legend().hide()

        self.orders_pie_chart_view = QChartView(self.orders_pie_chart)
        self.orders_pie_chart_view.setRenderHint(QPainter.Antialiasing)
        self.orders_pie_chart_view.chart().setBackgroundBrush(QBrush(QColor(50, 56, 66)))
        MainFunctions.created_widgets.append(self.orders_pie_chart_view)

        # Finances page
        self.finances_table = PyTableWidget()
        MainFunctions.created_widgets.append(self.finances_table)

        # Employees page
        self.employees_table = PyTableWidget()
        MainFunctions.created_widgets.append(self.employees_table)

        self.orders_splitter = QSplitter(Qt.Orientation.Horizontal)
        self.orders_splitter.addWidget(self.orders_table)
        self.orders_splitter.addWidget(self.orders_pie_chart_view)

        self.orders_splitter.setSizes([300, 300])
        self.orders_splitter.setHandleWidth(15)

        self.ui.orders_horizontal_layout.addWidget(self.orders_splitter)

        MainFunctions.created_widgets.append(self.orders_splitter)

        self.finances_splitter = QSplitter(Qt.Orientation.Horizontal)
        self.finances_splitter.addWidget(self.finances_table)

        self.finances_splitter.setSizes([300, 300])
        self.finances_splitter.setHandleWidth(15)

        self.ui.finances_horizontal_layout.addWidget(self.finances_splitter)
        MainFunctions.created_widgets.append(self.finances_splitter)

        self.employees_splitter = QSplitter(Qt.Orientation.Horizontal)
        self.employees_splitter.addWidget(self.employees_table)

        self.employees_splitter.setSizes([300, 300])
        self.employees_splitter.setHandleWidth(15)

        self.ui.employees_horizontal_layout.addWidget(self.employees_splitter)
        MainFunctions.created_widgets.append(self.employees_splitter)

        if self.data is not None and self.selected_file_paths is not None:
            selected_path = next((path for path in self.selected_file_paths if 'kelowna' in path.lower()), None)
            selected_path_square = next((path for path in self.selected_file_paths if 'item' in path.lower()), None)
            selected_path_employees = next((path for path in self.selected_file_paths if 'shifts' in path.lower()), None)
            selected_path_prep = next((path for path in self.selected_file_paths if 'pages' in path.lower()), None)
            selected_path_qbo = next((path for path in self.selected_file_paths if 'inc' in path.lower()), None)

            if selected_path is not None and selected_path_prep is not None:
                MainFunctions.fill_orders_page(self, selected_path, selected_path_prep, compare=False)
                MainFunctions.fill_finances_page(self, selected_path, compare=False)

            if selected_path_square is not None and selected_path_qbo is not None:
                MainFunctions.fill_square_finances(self, selected_path_square, selected_path_qbo, compare=False)

            if selected_path_employees is not None:
                MainFunctions.fill_employees(self, selected_path_employees, compare=False)

        if self.data is not None and self.compare_data is not None and self.selected_file_compare_paths is not None:
            selected_path_compare = next((path for path in self.selected_file_compare_paths if 'kelowna' in path.lower()), None)
            selected_path_prep_compare = next((path for path in self.selected_file_compare_paths if 'pages' in path.lower()), None)
            selected_path_square_compare = next((path for path in self.selected_file_compare_paths if 'item' in path.lower()), None)
            selected_path_employees_compare = next((path for path in self.selected_file_compare_paths if 'shifts' in path.lower()), None)
            selected_path_qbo_compare = next((path for path in self.selected_file_compare_paths if 'inc' in path.lower()), None)

            if selected_path_compare is not None and selected_path_prep_compare is not None:
                MainFunctions.fill_orders_page(self, selected_path_compare, selected_path_prep_compare, compare=True)
                MainFunctions.fill_finances_page(self, selected_path_compare, compare=True)

            if selected_path_square_compare is not None and selected_path_qbo_compare is not None:
                MainFunctions.fill_square_finances(self, selected_path_square_compare, selected_path_qbo_compare, compare=True)

            if selected_path_employees_compare is not None:
                MainFunctions.fill_employees(self, selected_path_employees_compare, compare=True)

    def fill_orders_page(self, selected_path, selected_path_prep, compare=False):
        if not compare:
            data_dict = self.create_orders_data_dict(self.data[selected_path], self.data[selected_path_prep])

            self.create_table_items(data_dict, self.orders_table, [self.date_range_file, "Counts"], self.orders_pie_chart_series)

        elif compare:
            # make compare order table
            orders_table = PyTableWidget(color="#4DA6FF")

            self.orders_splitter.insertWidget(1, orders_table)

            MainFunctions.created_widgets.append(orders_table)

            compare_data_dict = self.create_orders_compare_data_dict(self.compare_data[selected_path], self.compare_data[selected_path_prep])
            self.create_table_items(compare_data_dict, orders_table, [self.date_range_compare_file, 'Compare Counts'],
                                    compare_sum=False)

            # make compare sum order table
            orders_table_sum = PyTableWidget(color="#4DA6FF")

            MainFunctions.created_widgets.append(orders_table_sum)

            compare_data_dict_sum = self.create_orders_sum_compare_data_dict()
            self.create_table_items(compare_data_dict_sum, orders_table_sum, [self.date_range_compare_file, 'Sum Counts'],
                                    compare_sum=True)

            created_splitter = self.add_splitter(Qt.Orientation.Vertical, self.ui.orders_horizontal_layout,
                                                 [self.orders_splitter, orders_table_sum])

            MainFunctions.created_widgets.append(created_splitter)

    def fill_finances_page(self, selected_path, compare=False):
        if not compare:
            finances_data_dict = self.create_finances_data_dict(self.data[selected_path])

            self.create_table_items(finances_data_dict, self.finances_table, [self.date_range_file, "Figures"])

        elif compare:
            self.finances_bar_graph = QChart()
            self.finances_bar_graph.setBackgroundBrush(QBrush(QColor(50, 56, 66)))
            self.finances_bar_graph.legend().hide()

            finances_bar_series = QBarSeries()
            finances_bar_set_1 = QBarSet("Compare Data")
            finances_bar_set_2 = QBarSet("Data")
            finances_bar_set_1.setColor(QColor(77, 166, 255))
            finances_bar_set_2.setColor(QColor(255, 255, 255))

            self.finances_compare_table = PyTableWidget(color="#4DA6FF")

            self.finances_splitter.insertWidget(1, self.finances_compare_table)

            MainFunctions.created_widgets.append(self.finances_compare_table)

            finances_compare_data_dict = self.create_finances_data_dict_compare(self.compare_data[selected_path])

            self.create_table_items(finances_compare_data_dict, self.finances_compare_table, [self.date_range_compare_file,
                                                                               "Compare Figures"], bar_set=finances_bar_set_1)

            for item_bar_graph, additional_item_graph in self.data_dict_finances.items():
                finances_bar_set_2.append(float(additional_item_graph.replace('$', '').replace(',', '')))

            finances_bar_series.append(finances_bar_set_1)
            finances_bar_series.append(finances_bar_set_2)
            self.finances_bar_graph.addSeries(finances_bar_series)

            self.finances_axis_x = QBarCategoryAxis()
            self.finances_axis_x.append(self.compare_data_dict_finances.keys())
            self.finances_axis_x.setLabelsColor(QColor(255, 255, 255))
            self.finances_bar_graph.setAxisX(self.finances_axis_x, finances_bar_series)

            self.finances_bar_graph_view = QChartView(self.finances_bar_graph)
            self.finances_bar_graph_view.setRenderHint(QPainter.RenderHint.Antialiasing)
            self.finances_bar_graph_view.setGeometry(50, 50, 700, 300)

            self.finances_splitter.insertWidget(2, self.finances_bar_graph_view)

            MainFunctions.created_widgets.append(self.finances_bar_graph_view)

            self.finances_table_compare_sum = PyTableWidget(color="#4DA6FF")

            MainFunctions.created_widgets.append(self.finances_table_compare_sum)

            finances_compare_data_dict_sum = self.create_finances_data_dict_compare_sum()

            self.create_table_items(finances_compare_data_dict_sum, self.finances_table_compare_sum,
                                    [self.date_range_compare_file, "Sum Figures"], compare_sum_finances=True)

            created_splitter = self.add_splitter(Qt.Orientation.Vertical, self.ui.finances_horizontal_layout,
                                                 [self.finances_splitter, self.finances_table_compare_sum])

            MainFunctions.created_widgets.append(created_splitter)

    def fill_square_finances(self, selected_path, selected_qbo_path, compare=False):
        wb = openpyxl.load_workbook(selected_qbo_path, data_only=False)
        sheet = wb.active

        formulas_dict = {}

        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            for col_idx, cell_value in enumerate(row, start=1):
                if isinstance(cell_value, str) and cell_value.startswith('='):
                    cell_coord = openpyxl.utils.get_column_letter(col_idx) + str(row_idx)
                    formulas_dict[cell_coord] = cell_value[1:]  # Remove the '=' sign

        if not compare:
            # do the original table
            finances_square_data_dict = self.create_finances_square_data_dict(self.data[selected_path],
                                                                              formulas_dict)

            self.create_table_items(finances_square_data_dict, self.finances_table,
                                    [self.date_range_file,
                                     "Figures"])

        elif compare:
            bar_series = QBarSeries()
            finances_bar_set_1 = QBarSet("Compare Data")
            finances_bar_set_2 = QBarSet("Data")
            finances_bar_set_1.setColor(QColor(77, 166, 255))
            finances_bar_set_2.setColor(QColor(255, 255, 255))

            # do the compare table
            finances_square_data_dict_compare = self.create_finances_square_data_dict_compare(
                self.compare_data[selected_path], formulas_dict)

            self.create_table_items(finances_square_data_dict_compare, self.finances_compare_table,
                                    [self.date_range_compare_file,
                                     "Compare Figures"], bar_set=finances_bar_set_1)

            for item_bar_graph, additional_item_graph in self.data_dict_finances_square.items():
                finances_bar_set_2.append(
                    float(additional_item_graph.replace('$', '').replace(',', '')))

            bar_series.append(finances_bar_set_1)
            bar_series.append(finances_bar_set_2)

            self.finances_bar_graph.removeAllSeries()
            self.finances_bar_graph.addSeries(bar_series)
            self.finances_axis_x.clear()

            self.finances_axis_x.append(self.data_dict_finances_square.keys())
            self.finances_axis_x.setLabelsColor(QColor(255, 255, 255))
            self.finances_bar_graph.setAxisX(self.finances_axis_x, bar_series)

            # sums
            finances_compare_data_dict_sum = self.create_finances_square_data_dict_compare_sum()

            self.create_table_items(finances_compare_data_dict_sum, self.finances_table_compare_sum,
                                    [self.date_range_compare_file, "Sum Figures"],
                                    compare_sum_finances=True)

        wb.close()

    def fill_employees(self, selected_path, compare=False):
        if not compare:
            employees_data_dict = self.create_employees_data_dict(self.data[selected_path])

            self.create_table_items(employees_data_dict, self.employees_table,
                                    [self.date_range_file,
                                     "Hours"])
        elif compare:
            bar_graph = QChart()
            bar_graph.setBackgroundBrush(QBrush(QColor(50, 56, 66)))
            bar_graph.legend().hide()

            bar_series = QBarSeries()
            bar_set_1 = QBarSet("Compare Data")
            bar_set_2 = QBarSet("Data")

            bar_set_1.setColor(QColor(77, 166, 255))
            bar_set_2.setColor(QColor(255, 255, 255))

            compare_table = PyTableWidget(color="#4DA6FF")
            MainFunctions.created_widgets.append(compare_table)

            employees_data_dict_compare = self.create_employees_data_dict_compare(self.compare_data[selected_path])

            self.create_table_items(employees_data_dict_compare, compare_table,
                                    [self.date_range_compare_file,
                                     "Compare Hours"], bar_set=bar_set_1)

            for item_bar_graph, additional_item_graph in self.employee_data_dict.items():
                bar_set_2.append(float(additional_item_graph))

            self.employees_splitter.insertWidget(1, compare_table)

            bar_series.append(bar_set_1)
            bar_series.append(bar_set_2)
            bar_graph.addSeries(bar_series)

            axis_x = QBarCategoryAxis()
            axis_x.append(employees_data_dict_compare.keys())
            axis_x.setLabelsColor(QColor(255, 255, 255))
            bar_graph.setAxisX(axis_x, bar_series)

            bar_graph_view = QChartView(bar_graph)
            bar_graph_view.setRenderHint(QPainter.RenderHint.Antialiasing)
            bar_graph_view.setGeometry(50, 50, 700, 300)

            MainFunctions.created_widgets.append(bar_graph_view)

            self.employees_splitter.insertWidget(2, bar_graph_view)

            #sum table
            employees_table_sum = PyTableWidget(color="#4DA6FF")

            employees_data_dict_sum = self.create_employees_data_dict_compare_sum()
            self.create_table_items(employees_data_dict_sum, employees_table_sum,
                                    [self.date_range_compare_file, 'Sum Hours'],
                                    compare_sum=True)

            created_splitter = self.add_splitter(Qt.Orientation.Vertical, self.ui.employees_horizontal_layout,
                                                 [self.employees_splitter, employees_table_sum])

            MainFunctions.created_widgets.append(created_splitter)
